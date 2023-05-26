import openpyxl

def find_tables_with_pattern(filename, start_letter, end_letter):
    wb = openpyxl.load_workbook(filename)
    tables = []

    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        max_row = sheet.max_row
        max_col = sheet.max_column

        # Iterate over each cell in the sheet
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=row, column=col).value

                # Check if the cell value matches the pattern
                if (
                    cell_value
                    and isinstance(cell_value, str)
                    and cell_value.startswith(start_letter)
                    and cell_value.endswith(end_letter)
                ):
                    table_name = sheetname + "!" + sheet.cell(row=1, column=col).value
                    tables.append(table_name)

    return tables

# Example usage
filename = "example.xlsx"
start_letter = "A"
end_letter = "G"
tables = find_tables_with_pattern(filename, start_letter, end_letter)

if tables:
    print("Tables found:")
    for table in tables:
        print(table)
else:
    print("No tables found.")

from openpyxl import load_workbook
from pydantic.dataclasses import dataclass
path = "file_manager/tests/artifacts/20230224_triton_VVB.xlsx"
sheet_name = "Plate 1 - Sheet1 (2)"
wb = load_workbook(filename = path)
ws = wb.get_sheet_by_name(sheet_name)
print(ws.dimensions)

@dataclass
class UmaTabela:
    x2 : int
origins : list[UmaTabela] = [] 


# remover tudo até results:

#####
for row in ws.iter_cols():
    for col in row:
        if col.value and type(col.value) == str and "Results" in col.value: 
            print(row, col.col_idx)

#ws.

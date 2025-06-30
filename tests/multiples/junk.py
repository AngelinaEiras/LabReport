import openpyxl
import pandas as pd
import re

# Example usage
filename = "/home/angelina/Desktop/LabReport/src/file_manager/tests/artifacts/20230301_PB_triton_2.xlsx"
sheet_name = "Sheet1"
#filename = "/home/angelina/Desktop/LabReport/src/file_manager/tests/artifacts/20230224_triton_VVB.xlsx"
#sheet_name = "Plate 1 - Sheet1 (2)"

wb = openpyxl.load_workbook(filename)
ws = wb[sheet_name] # ws = wb.active # access the working sheet ws = wb['vgsales']

# print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))

# list to store the sub-datasets
subdatasets = []

# flag to identify the start of a sub-dataset
start_flag = False

# iterate over the rows of the worksheet
for row in ws.iter_rows(values_only=True):
    # check if the first cell starts with 'A'
    if row[0] and str(row[0]).startswith('A'):
        # set the start flag to True
        start_flag = True
        # create a new list to store the sub-dataset
        subdataset = [row]
    # check if the first cell starts with 'H'
    elif row[0] and str(row[0]).startswith('H'):
        # set the start flag to False
        start_flag = False
        # add the last row to the sub-dataset
        subdataset.append(row)
        # add the sub-dataset to the list
        subdatasets.append(subdataset)
    # add the row to the sub-dataset if the start flag is True
    elif start_flag:
        subdataset.append(row)

# iterate over the sub-datasets and print them
for i, subdataset in enumerate(subdatasets):
    print(f'Sub-dataset {i+1}:')
    for row in subdataset:
        print(row)
    print('\n')



def name_dataset (subdatasets):
    datasets = {}
    for i, subdataset in enumerate(subdatasets):
        name = input(f'Nome of dataset {i+1} ')
        datasets.update({name: subdataset})
    return datasets
# bo = name_dataset(subdatasets)



# Loop to select rows and columns
while True:
    # Ask the user which subdataset they want to work with
    which_subset = input('Enter the index for which subdataset you want to select (or enter "done" to exit): ')
    
    # Check if the user wants to exit
    if which_subset.lower() == 'done':
        print('Exiting...')
        break
        
    # Convert the user input to an integer
    subset_index = int(which_subset)
    
    # Initialize selected_rows and selected_cols variables to None
    selected_rows = None
    selected_cols = None
    
    # Loop to select rows or columns
    while True:
        # Ask the user if they want to select rows or columns
        selection = input('Do you want to select rows or columns? (Enter "rows" or "cols", or enter "done" to exit) ')
        
        # Check if the user wants to exit
        if selection.lower() == 'done':
            break
            
        elif selection.lower() == 'rows':
            # Ask the user which rows they want to select
            row_indices = input('Enter the row indices you want to select (separated by commas): ')
            selected_rows = [int(i) for i in row_indices.split(',')]

        elif selection.lower() == 'cols':
            # Ask the user which columns they want to select
            col_indices = input('Enter the column indices you want to select (separated by commas): ')
            selected_cols = [int(i) for i in col_indices.split(',')]
            
        else:
            # If the user enters an invalid input, break out of the loop
            print('Invalid input.')
            break
        
        # Ask the user if they want to select more rows or columns
        more_selection = input('Do you want to select more rows or columns? (Enter "yes" or "no") ')
        
        if more_selection.lower() == 'no':
            sub_df = subdataset.iloc[selected_rows, selected_cols]
            print(f'Sub-dataset {subset_index}:\n', sub_df.to_string(index=False), '\n\n')
            break


# TODO: melhorar a usabilidade da coisa, nomeadamente deixar alguns valores by default
# TODO: nomear os datasets, passar info


# Calculate the median
median = ws.median(sub_df)

# Calculate the mean
mean = ws.mean(sub_df)

# Calcule o desvio padrão
sd = ws.std(sub_df)

# Calcule o coeficiente de variação
cv = sd / ws.mean(sub_df) * 100
print(f"Median: {median}\nMean: {mean.item()}\nDesvio Padrão: {sd.item()}\nCoeficiente de Variação: {cv.item()}")






#########################################################################################


'''
import pandas as pd
from pandas import DataFrame
import csv
import numpy as np



# Load the Excel file into a dataframe
path = '/home/angelina/Desktop/LabReport/src/file_manager/tests/artifacts/20230301_PB_triton_2.xlsx'
df = pd.read_excel(path, header=None) #df[0,0] corresponde à cell A2 -> pq header!!!!

#print(df.iloc[3,0]) # df.iloc[rows, cols]

# list to store the sub-datasets
subdatasets = []

# flag to identify the start of a sub-dataset
start_flag = False

# iterate over the rows of the dataframe
for index, row in df.iterrows():
    # check if the first column starts with 'A'
    if str(row[0]).startswith('A'):
        # set the start flag to True
        start_flag = True
        # create a new dataframe to store the sub-dataset
        subdataset = pd.DataFrame(columns=df.columns)
    # check if the first column starts with 'H'
    elif str(row[0]).startswith('H'):
        # set the start flag to False
        start_flag = False
        # add the last row to the sub-dataset
        subdataset = pd.concat([subdataset, row.to_frame().T])
        # add the sub-dataset to the list
        subdatasets.append(subdataset)
    # add the row to the sub-dataset if the start flag is True
    elif start_flag:
        subdataset = pd.concat([subdataset, row.to_frame().T])

# print the sub-datasets
for i, subdataset in enumerate(subdatasets):
    # name = input("nome do subdataset")
    # associar o i ao nome
    print(f'Sub-dataset {i+1}:\n', subdataset.to_string(index=False), '\n\n')


    

# TODO: aprimorar isto para que se torne funcional:
#   - divisão não só dos vários datasets contidos no mesmo excel, e aceder a todos
#       - seleção das diferentes experiências:
#          - (dif[reag], reagentes, etc, ou mesmo tempos diferentes no mesmo "dataset")
# TODO: forma user friendly de se selecionar o que se quer de um excel !!!

# Loop to select rows and columns
while True:
    # Ask the user which subdataset they want to work with
    which_subset = input('Enter the index for which subdataset you want to select (or enter "done" to exit): ')
    
    # Check if the user wants to exit
    if which_subset.lower() == 'done':
        print('Exiting...')
        break
        
    # Convert the user input to an integer
    subset_index = int(which_subset)
    
    # Initialize selected_rows and selected_cols variables to None
    selected_rows = None
    selected_cols = None
    
    # Loop to select rows or columns
    while True:
        # Ask the user if they want to select rows or columns
        selection = input('Do you want to select rows or columns? (Enter "rows" or "cols", or enter "done" to exit) ')
        
        # Check if the user wants to exit
        if selection.lower() == 'done':
            break
            
        elif selection.lower() == 'rows':
            # Ask the user which rows they want to select
            row_indices = input('Enter the row indices you want to select (separated by commas): ')
            selected_rows = [int(i) for i in row_indices.split(',')]

        elif selection.lower() == 'cols':
            # Ask the user which columns they want to select
            col_indices = input('Enter the column indices you want to select (separated by commas): ')
            selected_cols = [int(i) for i in col_indices.split(',')]
            
        else:
            # If the user enters an invalid input, break out of the loop
            print('Invalid input.')
            break
        
        # Ask the user if they want to select more rows or columns
        more_selection = input('Do you want to select more rows or columns? (Enter "yes" or "no") ')
        
        if more_selection.lower() == 'no':
            sub_df = subdataset.iloc[selected_rows, selected_cols]
            print(f'Sub-dataset {subset_index}:\n', sub_df.to_string(index=False), '\n\n')
            break



# Calculate the median
median = np.median(sub_df)

# Calculate the mean
mean = np.mean(sub_df)

# Calcule o desvio padrão
sd = np.std(sub_df)

# Calcule o coeficiente de variação
cv = sd / np.mean(sub_df) * 100
print(f"Median: {median}\nMean: {mean.item()}\nDesvio Padrão: {sd.item()}\nCoeficiente de Variação: {cv.item()}")

'''


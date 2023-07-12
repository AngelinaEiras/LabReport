import openpyxl
from pydantic import dataclasses
import pandas as pd

@dataclasses.dataclass()
class LoadDataframeFromExcell:
    excell_path : str
    sheet_name: str = "Sheet1"

    def load_subdatasets(self) -> pd.DataFrame:
        wb = openpyxl.load_workbook(self.excell_path)
        ws = wb[self.sheet_name] # ws = wb.active # access the working sheet ws = wb['vgsales']

        # List to store the sub-datasets
        subdatasets = []

        # Flag to identify the start of a sub-dataset
        start_flag = False

        # iterate over the rows of the worksheet
        for row in ws.iter_rows(values_only=True):
            # check if the first cell starts with 'A'
            if row[0] and str(row[0]).startswith('A'):
                # set the start flag to True
                start_flag = True
                # create a new list to store the sub-dataset
                subdataset = [row]
            # check if the first cell starts with 'H'
            elif row[0] and str(row[0]).startswith('H'):
                # set the start flag to False
                start_flag = False
                # add the last row to the sub-dataset
                subdataset.append(row)
                # add the sub-dataset to the list
                subdatasets.append(subdataset)
            # add the row to the sub-dataset if the start flag is True
            elif start_flag:
                subdataset.append(row)

        subdatasets = pd.DataFrame(subdatasets[0])
        return subdatasets







# give me a report generator (as the above) following the next type of coding
# ie i a class to generate reports

# import openpyxl
# from pydantic import dataclasses
# import pandas as pd

# @dataclasses.dataclass()
# class LoadDataframeFromExcell:
#     excell_path : str
#     sheet_name: str = "Sheet1"

#     def load_subdatasets(self) -> pd.DataFrame:
#         wb = openpyxl.load_workbook(self.excell_path)
#         ws = wb[self.sheet_name] # ws = wb.active # access the working sheet ws = wb['vgsales']

#         # List to store the sub-datasets
#         subdatasets = []

#         # Flag to identify the start of a sub-dataset
#         start_flag = False

#         # iterate over the rows of the worksheet
#         for row in ws.iter_rows(values_only=True):
#             # check if the first cell starts with 'A'
#             if row[0] and str(row[0]).startswith('A'):
#                 # set the start flag to True
#                 start_flag = True
#                 # create a new list to store the sub-dataset
#                 subdataset = [row]
#             # check if the first cell starts with 'H'
#             elif row[0] and str(row[0]).startswith('H'):
#                 # set the start flag to False
#                 start_flag = False
#                 # add the last row to the sub-dataset
#                 subdataset.append(row)
#                 # add the sub-dataset to the list
#                 subdatasets.append(subdataset)
#             # add the row to the sub-dataset if the start flag is True
#             elif start_flag:
#                 subdataset.append(row)

#         subdatasets = pd.DataFrame(subdatasets[0])
#         return subdatasets